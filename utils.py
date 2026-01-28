import openpyxl
from typing import Iterable

def get_sheet_from_file(file: str):
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    sheet = wb.active
    return sheet

def get_anchors_coordinates(anchors, sheet):
    need = set(anchors)
    found = {}
    
    # Ограничиваем поиск первыми 100 колонками (или меньше), 
    # чтобы не сканировать пустые области справа
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_col=50, values_only=True), start=1):
        for col_idx, value in enumerate(row, start=1):
            if value in need:
                found[value] = (row_idx, col_idx)
                need.remove(value)
                
                if not need:
                    return found
    return found

def get_next_right_value(coord, sheet):
    row_idx, col_idx = coord
    
    # Берем только одну строку, начиная строго ПРАВЕЕ якоря
    # Не указываем max_col, чтобы не триггерить лишние расчеты
    row_iter = sheet.iter_rows(
        min_row=row_idx, 
        max_row=row_idx, 
        min_col=col_idx + 1, 
        values_only=True
    )
    
    # Извлекаем список значений (там будет только одна строка)
    row_values = next(row_iter, [])
    
    # Возвращаем первый элемент, который не None
    for value in row_values:
        if value is not None:
            return value
            
    return None

def get_last_value_in_row(coord, sheet):
    (row_idx, col_idx) = coord
    
    # Извлекаем только ОДНУ нужную строку
    # min_row и max_row одинаковые, чтобы не читать лишнее
    # min_col = col_idx + 1, чтобы начать поиск СРАЗУ после якоря
    target_row_iter = sheet.iter_rows(
        min_row=row_idx, 
        max_row=row_idx, 
        min_col=col_idx + 1, 
        values_only=True
    )
    
    # Берем первую (и единственную) строку из генератора
    row_values = next(target_row_iter, [])
    
    # Фильтруем None и берем последний элемент
    # (используем list comprehension или filter для скорости)
    data = [v for v in row_values if v is not None]
    
    return data[-1] if data else None